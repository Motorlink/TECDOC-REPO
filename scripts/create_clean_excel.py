#!/usr/bin/env python3
"""
Create clean Excel file from enriched CSV
Each field in its own column, properly formatted
"""

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_excel_from_csv(csv_file, excel_file):
    """Convert CSV to clean Excel with formatting"""
    
    print("\n" + "="*80)
    print("CREATING CLEAN EXCEL FILE")
    print("="*80 + "\n")
    
    # Read CSV
    print(f"Reading CSV: {csv_file}")
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        header = next(reader)
        rows = list(reader)
    
    print(f"  Columns: {len(header)}")
    print(f"  Rows: {len(rows)}")
    
    # Create workbook
    print(f"\nCreating Excel workbook...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Webisco Export"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    td_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    cell_alignment = Alignment(vertical="top", wrap_text=False)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Write header
    print("Writing header...")
    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Different color for TecDoc columns
        if col_name.startswith('TD_'):
            cell.fill = td_header_fill
        else:
            cell.fill = header_fill
    
    # Write data
    print("Writing data rows...")
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        if row_idx % 1000 == 0:
            print(f"  Written {row_idx-1} rows...")
    
    print(f"  Written {len(rows)} rows total")
    
    # Set column widths
    print("\nAdjusting column widths...")
    
    # Define specific widths for known columns
    column_widths = {
        'TD_ArticleNumber': 18,
        'TD_EAN': 15,
        'TD_MfrName': 15,
        'TD_GenericArticleDescription': 20,
        'TD_AssemblyGroupName': 20,
        'TD_OEMNumbers': 50,
        'TD_ArticleCriteria': 60,
        'TD_ArticleStatus': 12,
        'TD_ImageCount': 12,
        'TD_Bild_1_URL': 70,
        'TD_Bild_2_URL': 70,
        'TD_Bild_3_URL': 70,
        'TD_Bild_4_URL': 70,
        'TD_Bild_5_URL': 70,
        'TD_Bild_6_URL': 70,
        'TD_Bild_7_URL': 70,
        'TD_Bild_8_URL': 70,
        'TD_Bild_9_URL': 70,
        'TD_Bild_10_URL': 70,
    }
    
    for col_idx, col_name in enumerate(header, 1):
        if col_name in column_widths:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = column_widths[col_name]
        else:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
    
    # Freeze first row
    ws.freeze_panes = "A2"
    
    # Save workbook
    print(f"\nSaving Excel file: {excel_file}")
    wb.save(excel_file)
    
    print(f"\n{'='*80}")
    print("EXCEL FILE CREATED SUCCESSFULLY")
    print("="*80)
    print(f"File: {excel_file}")
    print(f"Columns: {len(header)}")
    print(f"Rows: {len(rows)}")
    print("="*80 + "\n")

if __name__ == "__main__":
    csv_file = "/home/ubuntu/Webisco_FINAL_Enriched_20251219_045315.csv"
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = f"/home/ubuntu/Webisco_FINAL_Export_{timestamp}.xlsx"
    
    create_excel_from_csv(csv_file, excel_file)
